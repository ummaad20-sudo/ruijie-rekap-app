from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.graphics import Color, Rectangle
from kivy.core.clipboard import Clipboard
from kivy.utils import get_color_from_hex
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime


class RekapApp(App):

    def format_rupiah(self, angka):
        try:
            return "Rp {:,}".format(int(angka)).replace(",", ".")
        except:
            return "Rp 0"

    def build(self):
        root = BoxLayout(orientation='vertical', padding=15, spacing=10)

        with root.canvas.before:
            Color(rgba=get_color_from_hex("#263238"))
            self.bg = Rectangle(size=root.size, pos=root.pos)

        root.bind(size=self.update_bg, pos=self.update_bg)

        title = Label(
            text="[b]REKAP DATA RUIJIE PRO[/b]",
            markup=True,
            font_size=22,
            size_hint=(1, 0.08),
            color=(1, 1, 1, 1)
        )
        root.add_widget(title)

        self.result_label = Label(
            text="Silakan pilih file Excel...",
            markup=True,
            size_hint_y=None,
            color=(1, 1, 1, 1),
            halign="center",
            valign="top"
        )

        self.result_label.bind(texture_size=self.result_label.setter('size'))
        self.result_label.bind(
            width=lambda s, w: s.setter('text_size')(s, (w - 20, None))
        )

        scroll = ScrollView(size_hint=(1, 0.70))
        scroll.add_widget(self.result_label)
        root.add_widget(scroll)

        tombol_box = BoxLayout(size_hint=(1, 0.10), spacing=20)

        btn_file = Button(
            text="📂 PILIH FILE EXCEL",
            background_normal="",
            background_color=(1, 0.5, 0.1, 1)
        )

        btn_share = Button(
            text="📋 COPY HASIL",
            background_normal="",
            background_color=(0.1, 0.4, 0.9, 1)
        )

        btn_file.bind(on_press=self.buka_file)
        btn_share.bind(on_press=self.share_hasil)

        tombol_box.add_widget(btn_file)
        tombol_box.add_widget(btn_share)
        root.add_widget(tombol_box)

        self.notif = Label(
            text="",
            size_hint=(1, 0.04),
            color=(1, 0.9, 0.3, 1)
        )
        root.add_widget(self.notif)

        footer = Label(
            text="Creator by JUN.AI © 2026",
            size_hint=(1, 0.05),
            font_size=12,
            color=(0.7, 0.7, 0.7, 1)
        )
        root.add_widget(footer)

        return root

    def update_bg(self, instance, value):
        self.bg.size = instance.size
        self.bg.pos = instance.pos

    def buka_file(self, instance):
        chooser = FileChooserListView(
            path="/storage/emulated/0",
            filters=["*.xlsx"]
        )

        popup = Popup(
            title="Pilih File Excel",
            content=chooser,
            size_hint=(0.95, 0.95)
        )

        chooser.bind(
            on_submit=lambda x, selection, touch:
            self.proses_file(selection, popup)
        )

        popup.open()

    def proses_file(self, selection, popup):
        if not selection:
            return

        file_path = selection[0]
        popup.dismiss()

        try:
            wb = load_workbook(file_path)
            sheet = wb.active
        except:
            self.result_label.text = "[color=ff4444]Gagal membuka file Excel[/color]"
            return

        rekap_detail = defaultdict(lambda: {"jumlah": 0, "total": 0})
        rekap_tanggal = defaultdict(int)

        header = [cell.value for cell in sheet[1]]

        try:
            kolom_grup = header.index("Grup pengguna")
            kolom_harga = header.index("Harga")
            kolom_tanggal = header.index("Diaktifkan di")
        except:
            self.result_label.text = "[color=ff4444]Header Excel tidak sesuai[/color]"
            return

        for row in sheet.iter_rows(min_row=2, values_only=True):
            grup = row[kolom_grup]
            harga = row[kolom_harga]
            tanggal_full = row[kolom_tanggal]

            if grup and harga and tanggal_full:
                if isinstance(tanggal_full, datetime):
                    tanggal = tanggal_full.strftime("%Y/%m/%d")
                else:
                    tanggal = str(tanggal_full).split(" ")[0]

                try:
                    harga_int = int(harga)
                except:
                    continue

                key = (tanggal, grup)
                rekap_detail[key]["jumlah"] += 1
                rekap_detail[key]["total"] += harga_int
                rekap_tanggal[tanggal] += harga_int

        hasil = "[b]===== HASIL REKAP =====[/b]\n\n"
        tanggal_terakhir = None

        for (tanggal, grup), data in sorted(rekap_detail.items()):
            if tanggal_terakhir and tanggal != tanggal_terakhir:
                hasil += f"[color=00ffcc][b]TOTAL {self.format_rupiah(rekap_tanggal[tanggal_terakhir])}[/b][/color]\n"
                hasil += "-" * 40 + "\n\n"

            if tanggal != tanggal_terakhir:
                hasil += f"[color=FFA726][b]Tanggal : {tanggal}[/b][/color]\n"
                hasil += "-" * 40 + "\n"

            hasil += f"Grup   : {grup}\n"
            hasil += f"Jumlah : {data['jumlah']}\n"
            hasil += f"[b]Total  : {self.format_rupiah(data['total'])}[/b]\n"
            hasil += "-" * 40 + "\n"

            tanggal_terakhir = tanggal

        if tanggal_terakhir:
            hasil += f"[color=00ffcc][b]TOTAL {self.format_rupiah(rekap_tanggal[tanggal_terakhir])}[/b][/color]\n"

        self.result_label.text = hasil
        self.notif.text = "✅ File berhasil diproses"

    def share_hasil(self, instance):
        if not self.result_label.text.strip():
            self.notif.text = "⚠ Tidak ada hasil"
            return

        Clipboard.copy(self.result_label.text)
        self.notif.text = "📋 Hasil disalin ke clipboard"


if __name__ == "__main__":
    RekapApp().run()
